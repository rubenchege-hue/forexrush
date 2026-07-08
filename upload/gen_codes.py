import random, string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def make_code(prefix="COMP"):
    part = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
    return f"{prefix}-{part[:4]}-{part[4:]}"

random.seed(42)

wb = Workbook()
ws = wb.active
ws.title = "Access Codes"

headers = ["Client Name", "Email", "Access Code", "Issue Date", "Expiry Date (30 days)", "Days Remaining", "Status", "Redeemed (Y/N)"]
header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
header_font = Font(bold=True, color="FFFFFF", name="Arial")
thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

ws.freeze_panes = "A2"

sample_names = [f"Client {i}" for i in range(1, 16)]
start_row = 2
for i, name in enumerate(sample_names):
    r = start_row + i
    email_cell = f"B{r}"
    code = make_code()
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=f"client{i+1}@example.com")
    code_cell = ws.cell(row=r, column=3, value=code)
    code_cell.font = Font(name="Consolas", bold=True, color="0000FF")
    issue_cell = ws.cell(row=r, column=4, value="=TODAY()")
    issue_cell.number_format = "yyyy-mm-dd"
    expiry_cell = ws.cell(row=r, column=5, value=f"=D{r}+30")
    expiry_cell.number_format = "yyyy-mm-dd"
    days_left = ws.cell(row=r, column=6, value=f"=MAX(E{r}-TODAY(),0)")
    status_cell = ws.cell(row=r, column=7, value=f'=IF(TODAY()>E{r},"Expired","Active")')
    ws.cell(row=r, column=8, value="N")
    for col in range(1, 9):
        ws.cell(row=r, column=col).border = border
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="left")

widths = [20, 26, 20, 16, 20, 15, 12, 15]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Instructions sheet
inst = wb.create_sheet("How To Use")
inst_lines = [
    ("Trading Competition Access Code Tracker", True, 14),
    ("", False, 11),
    ("1. Add each paying client's name and email in columns A and B.", False, 11),
    ("2. Generate a new unique code (see 'Generate More Codes' below) and paste into column C.", False, 11),
    ("3. Issue Date auto-fills to today when you open the file; Expiry Date auto-calculates as Issue Date + 30 days.", False, 11),
    ("4. Status automatically switches from Active to Expired once the 30-day window passes.", False, 11),
    ("5. Mark column H 'Y' once a client has redeemed/claimed their result.", False, 11),
    ("", False, 11),
    ("Generate More Codes:", True, 12),
    ("Run the attached gen_codes.py script (or ask Claude) any time you need a fresh batch of unique codes.", False, 11),
    ("", False, 11),
    ("Note: This sheet only tracks codes and expiry - it does not connect to any trading platform.", False, 11),
    ("You'll still need to manually grant/revoke access on your platform when codes are issued or expire.", False, 11),
]
for i, (text, bold, size) in enumerate(inst_lines, 1):
    c = inst.cell(row=i, column=1, value=text)
    c.font = Font(bold=bold, size=size, name="Arial")
inst.column_dimensions['A'].width = 100

wb.save("trading_competition_codes.xlsx")
print("saved")
